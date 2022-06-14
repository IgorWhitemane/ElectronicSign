import re
import shutil
import time
import openpyxl
import os
import datetime
import configparser

from PyQt5.QtWidgets import QMessageBox
from PyQt5 import QtCore, QtWidgets
from PyQt5.QtGui import QIcon

from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By

from docxtpl import DocxTemplate


def initial_filling(self):
    try:
        self.ui.all_people.clear()
        config = configparser.ConfigParser()
        config.read("doc\\settings.ini", encoding="utf-8")
        wb = openpyxl.reader.excel.load_workbook(config["Path"]["path_db_all"])
        row_max = wb.active.max_row
        for i in range(row_max - 1):
            full_name = str(wb.active['B' + str(i + 2)].value)
            gender = str(wb.active['C' + str(i + 2)].value)
            position = str(wb.active['D' + str(i + 2)].value)
            department = str(wb.active['H' + str(i + 2)].value)
            if not department or department == "None" or department is None:
                department = str(wb.active['G' + str(i + 2)].value)
                if not department or department == "None" or department is None:
                    department = str(wb.active['F' + str(i + 2)].value)
                    if not department or department == "None" or department is None:
                        department = str(wb.active['E' + str(i + 2)].value)
            email = str(wb.active['I' + str(i + 2)].value)
            passport_ser_num = str(wb.active['J' + str(i + 2)].value)
            passport_issued = str(wb.active['L' + str(i + 2)].value)
            passport_kode = str(wb.active['M' + str(i + 2)].value)
            passport_date = str(wb.active['N' + str(i + 2)].value)
            date_of_birth = str(wb.active['O' + str(i + 2)].value)
            city_of_birth = str(wb.active['P' + str(i + 2)].value)
            snils = str(wb.active['Q' + str(i + 2)].value)
            inn = str(wb.active['R' + str(i + 2)].value)
            number_phone = str(wb.active['S' + str(i + 2)].value)

            QtWidgets.QTreeWidgetItem(self.ui.all_people)

            self.ui.all_people.topLevelItem(i).setText(0, QtCore.QCoreApplication.translate("MainWindow",
                                                                                            full_name))
            self.ui.all_people.topLevelItem(i).setText(1, QtCore.QCoreApplication.translate("MainWindow",
                                                                                            gender))
            self.ui.all_people.topLevelItem(i).setText(2, QtCore.QCoreApplication.translate("MainWindow",
                                                                                            position))
            self.ui.all_people.topLevelItem(i).setText(3, QtCore.QCoreApplication.translate("MainWindow",
                                                                                            department))
            self.ui.all_people.topLevelItem(i).setText(4, QtCore.QCoreApplication.translate("MainWindow",
                                                                                            email))
            self.ui.all_people.topLevelItem(i).setText(5, QtCore.QCoreApplication.translate("MainWindow",
                                                                                            passport_ser_num))
            self.ui.all_people.topLevelItem(i).setText(6, QtCore.QCoreApplication.translate("MainWindow",
                                                                                            passport_issued))
            self.ui.all_people.topLevelItem(i).setText(7, QtCore.QCoreApplication.translate("MainWindow",
                                                                                            passport_kode))
            self.ui.all_people.topLevelItem(i).setText(8, QtCore.QCoreApplication.translate("MainWindow",
                                                                                            passport_date))
            self.ui.all_people.topLevelItem(i).setText(9, QtCore.QCoreApplication.translate("MainWindow",
                                                                                            date_of_birth))
            self.ui.all_people.topLevelItem(i).setText(10, QtCore.QCoreApplication.translate("MainWindow",
                                                                                             city_of_birth))
            self.ui.all_people.topLevelItem(i).setText(11, QtCore.QCoreApplication.translate("MainWindow",
                                                                                             snils))
            self.ui.all_people.topLevelItem(i).setText(12, QtCore.QCoreApplication.translate("MainWindow",
                                                                                             inn))
            self.ui.all_people.topLevelItem(i).setText(13, QtCore.QCoreApplication.translate("MainWindow",
                                                                                             number_phone))
    except (Exception,):
        mistake = 'Ошибка в пути к файлу базы данных'
        exception(mistake)


def on_btn_search_clicked(self):
    self.ui.all_people.clear()
    item = self.ui.search.text()
    if not item:
        initial_filling(self)
    else:
        search_people(self, item)


def search_people(self, item):
    try:
        search_text = item.lower()
        config = configparser.ConfigParser()
        config.read("doc\\settings.ini", encoding="utf-8")
        wb = openpyxl.load_workbook(config["Path"]["path_db_all"])
        row_max = wb.active.max_row
        item_number = 0

        for i in range(row_max):
            data_from_cell = wb.active["B" + str(i + 1)].value
            data_from_cell = str(data_from_cell)
            data_from_cell = data_from_cell.lower()
            result_search = re.findall(search_text, data_from_cell)
            if len(result_search) > 0:
                result_row = i + 1

                full_name = str(wb.active['B' + str(result_row)].value)
                gender = str(wb.active['C' + str(result_row)].value)
                position = str(wb.active['D' + str(result_row)].value)
                department = str(wb.active['H' + str(result_row)].value)
                if not department or department == "None" or department is None:
                    department = str(wb.active['G' + str(result_row)].value)
                    if not department or department == "None" or department is None:
                        department = str(wb.active['F' + str(result_row)].value)
                        if not department or department == "None" or department is None:
                            department = str(wb.active['E' + str(result_row)].value)
                email = str(wb.active['I' + str(result_row)].value)
                passport_ser_num = str(wb.active['J' + str(result_row)].value)
                passport_issued = str(wb.active['L' + str(result_row)].value)
                passport_kode = str(wb.active['M' + str(result_row)].value)
                passport_date = str(wb.active['N' + str(result_row)].value)
                date_of_birth = str(wb.active['O' + str(result_row)].value)
                city_of_birth = str(wb.active['P' + str(result_row)].value)
                snils = str(wb.active['Q' + str(result_row)].value)
                inn = str(wb.active['R' + str(result_row)].value)
                number_phone = str(wb.active['S' + str(result_row)].value)
                QtWidgets.QTreeWidgetItem(self.ui.all_people)

                self.ui.all_people.topLevelItem(item_number).setText(0, QtCore.QCoreApplication.translate("MainWindow",
                                                                                                          full_name))
                self.ui.all_people.topLevelItem(item_number).setText(1, QtCore.QCoreApplication.translate("MainWindow",
                                                                                                          gender))
                self.ui.all_people.topLevelItem(item_number).setText(2, QtCore.QCoreApplication.translate("MainWindow",
                                                                                                          position))
                self.ui.all_people.topLevelItem(item_number).setText(3, QtCore.QCoreApplication.translate("MainWindow",
                                                                                                          department))
                self.ui.all_people.topLevelItem(item_number).setText(4, QtCore.QCoreApplication.translate("MainWindow",
                                                                                                          email))
                self.ui.all_people.topLevelItem(item_number).setText(5, QtCore.QCoreApplication.translate("MainWindow",
                                                                                                          passport_ser_num))
                self.ui.all_people.topLevelItem(item_number).setText(6, QtCore.QCoreApplication.translate("MainWindow",
                                                                                                          passport_issued))
                self.ui.all_people.topLevelItem(item_number).setText(7, QtCore.QCoreApplication.translate("MainWindow",
                                                                                                          passport_kode))
                self.ui.all_people.topLevelItem(item_number).setText(8, QtCore.QCoreApplication.translate("MainWindow",
                                                                                                          passport_date))
                self.ui.all_people.topLevelItem(item_number).setText(9, QtCore.QCoreApplication.translate("MainWindow",
                                                                                                          date_of_birth))
                self.ui.all_people.topLevelItem(item_number).setText(10, QtCore.QCoreApplication.translate("MainWindow",
                                                                                                           city_of_birth))
                self.ui.all_people.topLevelItem(item_number).setText(11, QtCore.QCoreApplication.translate("MainWindow",
                                                                                                           snils))
                self.ui.all_people.topLevelItem(item_number).setText(12, QtCore.QCoreApplication.translate("MainWindow",
                                                                                                           inn))
                self.ui.all_people.topLevelItem(item_number).setText(13, QtCore.QCoreApplication.translate("MainWindow",
                                                                                                           number_phone))
                item_number += 1

    except (Exception,):
        mistake = 'Недопустимые символы в поиске!'
        exception(mistake)


def on_item_clicked(self):
    try:
        status = None
        result_column = None
        item = self.ui.all_people.currentItem()
        full_name = item.text(0)
        self.ui.result.setText(f"{full_name}")
        search_text = full_name.lower()
        config = configparser.ConfigParser()
        config.read("doc\\settings.ini", encoding="utf-8")
        wb = openpyxl.load_workbook(config["Path"]["path_reestr_es"])
        row_max = wb['Реестр выпуска с Сибинтек'].max_row

        for i in range(row_max):
            data_from_cell = wb['Реестр выпуска с Сибинтек']["A" + str(i + 1)].value
            data_from_cell = str(data_from_cell)
            data_from_cell = data_from_cell.lower()
            result_search = re.findall(search_text, data_from_cell)
            if len(result_search) > 0:
                result_row = i + 1
                for j in range(10, 3, -1):
                    status = wb.get_sheet_by_name('Реестр выпуска с Сибинтек').cell(row=result_row, column=j).value
                    if status is not None:
                        result_column = j
                        break

        if result_column == 10:
            res = "ЭП записана:"
        elif result_column == 9:
            res = "Проверено:"
        elif result_column == 8:
            res = "Загружено на проверку:"
        elif result_column == 7:
            res = "Подписано у ГИ:"
        elif result_column == 6:
            res = "Подписано сотрудником:"
        elif result_column == 5:
            res = "Подготовлено заявление:"
        elif result_column == 4:
            res = "Документы получены:"
        else:
            res = "Документов нет."
            status = ""

        try:
            status = status.strftime('%d.%m.%Y')
        except (Exception,):
            pass
        self.ui.status.setText(f'{res} {status}')

    except (Exception,):
        mistake = 'Ошибка в выборе пользователя!'
        exception(mistake)


def config_write(page, row):
    config = configparser.ConfigParser()
    config.add_section('Page')
    config.read("doc\\settings.ini", encoding="utf-8")
    config.set('Page', 'page', str(page))
    config.set('Page', 'row', str(row))

    with open("doc\\settings.ini", 'w', encoding="utf-8") as f:
        config.write(f)


def on_btn_sender_clicked(self):
    options = Options()
    options.add_experimental_option("detach", True)
    options.add_argument("--start-maximized")
    item = self.ui.all_people.currentItem()
    name = item.text(0)
    config = configparser.RawConfigParser()
    config.read("doc\\settings.ini", encoding="utf-8")
    page = int(config["Page"]["page"])
    row = int(config["Page"]["row"])

    try:
        item = self.ui.all_people.currentItem()

        full_name = item.text(0)
        gender = item.text(1)
        position = item.text(2)
        department = item.text(3)
        passport_ser_num = item.text(5)
        passport_issued = item.text(6)
        passport_kode = item.text(7)
        passport_date = item.text(8)
        date_of_birth = item.text(9)
        city_of_birth = item.text(10)
        snils = item.text(11)
        inn = item.text(12)

        os.mkdir(f'{config["Path"]["path_for_user"]}{full_name}')
        os.mkdir(f'{config["Path"]["path_for_user"]}{full_name}\\сканы')

        try:
            # СТРАНИЦА АВТОРИЗАЦИИ
            driver = webdriver.Chrome(executable_path='driver/chromedriver.exe', options=options)
            url = f'{config["URL"]["url"]}{page}'
            driver.get(url)
            user = driver.find_element(By.XPATH, '//*[@id="loginform-username"]')
            user.send_keys(config["URL"]["login"])
            password = driver.find_element(By.XPATH, '//*[@id="loginform-password"]')
            password.send_keys(config["URL"]["pass"])
            element = driver.find_element(By.XPATH, '//*[@id="login-form"]/div[4]/div/button')
            element.click()

            # ПЕРВАЯ СТРАНИЦА
            row_number = f'//*[@id="w0"]/table/tbody/tr[{row}]/td[11]/button'
            row_number_two = f'//*[@id="w0"]/table/tbody/tr[{row}]/td[11]/a'
            try:
                wait = WebDriverWait(driver, 5)
                wait.until(EC.element_to_be_clickable((By.XPATH, row_number))).click()
                wait.until(EC.alert_is_present()).accept()
            except (Exception,):
                pass
            try:
                wait = WebDriverWait(driver, 5)
                wait.until(EC.element_to_be_clickable((By.XPATH, row_number_two))).click()
                wait.until(EC.alert_is_present()).accept()
            except (Exception,):
                pass

            # ВТОРАЯ СТРАНИЦА (ВВОД ДАННЫХ)
            driver.switch_to.window(driver.window_handles[1])

            wait = WebDriverWait(driver, 10)
            element = wait.until(EC.presence_of_element_located(
                (By.XPATH, '//*[@id="ModernCertificateCustomerForm_organizationShortName"]')))
            element.clear()
            element.send_keys('ООО "СЛАВНЕФТЬ-КРАСНОЯРСКНЕФТЕГАЗ"')

            element = driver.find_element(By.XPATH, '//*[@id="ModernCertificateCustomerForm_headLastName"]')
            element.clear()
            element.send_keys('Гребенюк')

            element = driver.find_element(By.XPATH, '//*[@id="ModernCertificateCustomerForm_headFirstName"]')
            element.clear()
            element.send_keys('Александр')

            element = driver.find_element(By.XPATH, '//*[@id="ModernCertificateCustomerForm_headMiddleName"]')
            element.clear()
            element.send_keys('Николаевич')

            element = driver.find_element(By.XPATH, '//*[@id="ModernCertificateCustomerForm_headPosition"]')
            element.clear()
            element.send_keys('Первый заместитель генерального директора по производству - главный инженер')

            driver.find_element(By.XPATH,
                                '//*[@id="ModernCertificateCustomerForm_authorityDocumentType"]/option[2]').click()

            element = driver.find_element(By.XPATH, '//*[@id="ModernCertificateCustomerForm_authorityDocumentDate"]')
            element.send_keys('10022022')

            element = driver.find_element(By.XPATH, '//*[@id="ModernCertificateCustomerForm_authorityDocumentNumber"]')
            element.send_keys('40')

            full_name = full_name.split()

            element = driver.find_element(By.XPATH, '//*[@id="ModernCertificateCustomerForm_ownerLastName"]')
            element.send_keys(full_name[0])

            element = driver.find_element(By.XPATH, '//*[@id="ModernCertificateCustomerForm_ownerFirstName"]')
            element.send_keys(full_name[1])

            element = driver.find_element(By.XPATH, '//*[@id="ModernCertificateCustomerForm_ownerMiddleName"]')
            element.send_keys(full_name[2])

            if gender == "мужской":
                driver.find_element(By.XPATH,
                                    '//*[@id="ModernCertificateCustomerForm_ownerGender"]/option[2]').click()
            else:
                driver.find_element(By.XPATH,
                                    '//*[@id="ModernCertificateCustomerForm_ownerGender"]/option[3]').click()

            element = driver.find_element(By.XPATH, '//*[@id="ModernCertificateCustomerForm_ownerINN"]')
            element.send_keys(inn)

            passport_ser_num = passport_ser_num.split()

            element = driver.find_element(By.XPATH, '//*[@id="ModernCertificateCustomerForm_passportSeries"]')
            element.send_keys(passport_ser_num[0], passport_ser_num[1])

            element = driver.find_element(By.XPATH, '//*[@id="ModernCertificateCustomerForm_passportNumber"]')
            element.send_keys(passport_ser_num[2])

            passport_date = passport_date.split()
            passport_date = passport_date[0]
            passport_date = passport_date.split("-")

            element = driver.find_element(By.XPATH, '//*[@id="ModernCertificateCustomerForm_passportDate"]')
            element.send_keys("".join(passport_date[::-1]))

            b = 0
            try:
                while b != passport_kode:
                    element = driver.find_element(By.XPATH,
                                                  '//*[@id="ModernCertificateCustomerForm_ownerPassportDeptCode"]')
                    element.clear()
                    element.send_keys(passport_kode)

                    wait = WebDriverWait(driver, 5)
                    a = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="deptSuggestionsBox"]/div[1]')))
                    b = a.text.split()
                    b = b[0]
                    a.click()
            except (Exception,):
                pass

            date_of_birth = date_of_birth.split()
            date_of_birth = date_of_birth[0]
            date_of_birth = date_of_birth.split("-")

            element = driver.find_element(By.XPATH, '//*[@id="ModernCertificateCustomerForm_ownerBirthDate"]')
            element.send_keys("".join(date_of_birth[::-1]))

            city_of_birth = city_of_birth.upper()

            element = driver.find_element(By.XPATH, '//*[@id="ModernCertificateCustomerForm_ownerBirthPlace"]')
            element.send_keys(city_of_birth)

            passport_issued = passport_issued.upper()

            element = driver.find_element(By.XPATH, '//*[@id="ModernCertificateCustomerForm_passportIssuer"]')
            element.clear()
            element.send_keys(passport_issued)

            element = driver.find_element(By.XPATH, '//*[@id="ModernCertificateCustomerForm_SNILS"]')
            element.send_keys(snils)

            element = driver.find_element(By.XPATH, '//*[@id="ModernCertificateCustomerForm_email"]')
            element.send_keys('ib@snkng.ru')

            element = driver.find_element(By.XPATH, '//*[@id="ModernCertificateCustomerForm_phone"]')
            element.send_keys('9130407388')

            element = driver.find_element(By.XPATH, '//*[@id="ModernCertificateCustomerForm_ownerPosition"]')
            element.send_keys(position)

            element = driver.find_element(By.XPATH, '//*[@id="ModernCertificateCustomerForm_departmentName"]')
            element.clear()
            element.send_keys(department)

            try:
                row += 1
                if row > 20:
                    page += 1
                    row = 1
                config_write(page, row)
            except (Exception, ):
                mistake = 'Не удалось изменить конфиг фаил'
                exception(mistake)

            write_log(item)
            open_rar(name)

            # ТРЕТЬЯ СТРАНИЦА (ЗАГРУЗКА ЗАЯВЛЕНИЯ В ПАПКУ К ПОЛЬЗОВАТЕЛЮ)
            try:
                wait = WebDriverWait(driver, 1000)
                wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="download-documents-next"]/a[2]'))).click()
                wait.until(EC.alert_is_present()).accept()
                time.sleep(4)
                driver.quit()
                move_file(name)
            except (Exception,):
                mistake = 'Заявление не скачалось'
                exception(mistake)
        except (Exception,):
            mistake = 'Возникла ошибка на сайте, при заполнении пользователя!'
            exception(mistake)

    except (Exception,):
        mistake = f'Пользователь {name} загружен ранее!'
        exception(mistake)


def write_log(item):
    full_name = item.text(0)
    date_now = datetime.datetime.today().strftime('%d.%m.%Y')

    f = open("doc\\log.txt", 'a')
    f.write(f'{full_name} ------- {date_now} \n')
    f.close()

    new_f = open("doc\\log_not_data.txt", 'a')
    new_f.write(f'{full_name}\n')
    new_f.close()


def open_log():
    os.startfile('doc\\log.txt')


def open_log_not_data():
    os.startfile('doc\\log_not_data.txt')


def open_rar(name):
    config = configparser.ConfigParser()
    config.read("doc\\settings.ini", encoding="utf-8")

    try:
        try:
            os.startfile(f'{config["Path"]["path_for_archives"]}{name}.7z')
            shutil.copy(f'{config["Path"]["path_for_archives"]}{name}.7z',
                        f'{config["Path"]["path_for_user"]}{name}')
        except (Exception,):
            pass

        try:
            os.startfile(f'{config["Path"]["path_for_archives"]}{name}.zip')
            shutil.copy(f'{config["Path"]["path_for_archives"]}{name}.zip',
                        f'{config["Path"]["path_for_user"]}{name}')
        except (Exception,):
            pass

        try:
            os.startfile(f'{config["Path"]["path_for_archives"]}{name}.rar')
            shutil.copy(f'{config["Path"]["path_for_archives"]}{name}.rar',
                        f'{config["Path"]["path_for_user"]}{name}')
        except (Exception,):
            pass
    except (Exception,):
        mistake = f'Не найдены документы на пользователя: {name}!'
        exception(mistake)


def move_file(name):
    config = configparser.ConfigParser()
    config.read("doc\\settings.ini", encoding="utf-8")
    try:
        while True:
            a = os.listdir(config["Path"]["path_downloads"])
            if a[1].endswith('.pdf'):
                shutil.move(f'{config["Path"]["path_downloads"]}\\{a[1]}',
                            f'{config["Path"]["path_for_user"]}\\{name}')
                break
    except (Exception,):
        mistake = 'Заявление не перенеслось, заберите из Download'
        exception(mistake)


def write_receipt(self):
    try:
        config = configparser.ConfigParser()
        config.read("doc\\settings.ini", encoding="utf-8")
        item = self.ui.all_people.currentItem()
        ser_num = self.ui.ser_num.text()
        full_name = item.text(0)
        position = item.text(2)
        department = item.text(3)
        number_token = search_token(full_name)
        doc = DocxTemplate("doc\\шаблон.docx")
        context = {'name': full_name, 'number': ser_num, 'position': position, 'posname': department,
                   'token': number_token}
        doc.render(context)
        doc.save(f'{config["Path"]["path_for_receipt"]}{full_name}.docx')
        reestr = 'Расписка за выдачу'
        result_row = 'Не требуется'
        successfully(result_row, full_name, reestr)
    except (Exception,):
        mistake = 'Не удалось создать расписку!'
        exception(mistake)


def search_token(full_name):
    config = configparser.ConfigParser()
    config.read("doc\\settings.ini", encoding="utf-8")
    search_text = full_name.lower()
    wb = openpyxl.load_workbook(config["Path"]["path_pkzi"])
    row_max = wb.active.max_row
    item_number = []
    for i in range(row_max):
        data_from_cell = wb.active["A" + str(i + 1)].value
        data_from_cell = str(data_from_cell)
        data_from_cell = data_from_cell.lower()
        result_search = re.findall(search_text, data_from_cell)
        if len(result_search) > 0:
            result_row = i + 1
            item_number.append(result_row)
    for i in item_number:
        num_token = str(wb.active['P' + str(i)].value)
        num_token_try = str(wb.active['Q' + str(i)].value)
        if num_token != "None" and (num_token_try == "None" or num_token_try == "Смена фамилии"):
            return num_token


def write_restr_one(self):
    try:
        reestr = None
        config = configparser.ConfigParser()
        config.read("doc\\settings.ini", encoding="utf-8")
        date_of_completion = datetime.datetime.today().strftime('%d.%m.%Y')
        item = self.ui.all_people.currentItem()
        full_name = item.text(0)
        search_text = full_name.lower()
        wb = openpyxl.load_workbook(config["Path"]["path_reestr_es"], read_only=False)
        row_max = wb.active.max_row
        for i in range(row_max):
            data_from_cell = wb.active["A" + str(i + 1)].value
            data_from_cell = str(data_from_cell)
            data_from_cell = data_from_cell.lower()
            result_search = re.findall(search_text, data_from_cell)
            if len(result_search) > 0:
                result_row = i + 1
                sheet = wb.get_sheet_by_name('Реестр выпуска с Сибинтек')
                sheet.cell(row=result_row, column=10).value = date_of_completion
                sheet.cell(row=result_row, column=11).value = date_of_completion
                sheet.cell(row=result_row, column=12).value = date_of_completion
                wb.save(config["Path"]["path_reestr_es"])
                reestr = 'Реестр выпуска ЭЦП'
                successfully(result_row, full_name, reestr)
        if reestr is None:
            mistake = f'Не удалось найти пользователя {full_name}'
            exception(mistake)
    except (Exception,):
        mistake = 'Не удалось отметить в реестре выпуска\nПроверьте что файл не кто не использует!'
        exception(mistake)


def write_restr_two(self):
    try:
        config = configparser.ConfigParser()
        config.read("doc\\settings.ini", encoding="utf-8")
        result_row = None
        item = self.ui.all_people.currentItem()
        full_name = item.text(0)
        search_text = full_name.lower()
        wb = openpyxl.load_workbook(config["Path"]["path_reestr_es_full"])
        row_max = wb['Реестр выпущенных ЭП'].max_row
        for i in range(row_max):
            data_from_cell = wb['Реестр выпущенных ЭП']["A" + str(i + 1)].value
            data_from_cell_two = str(data_from_cell)
            data_from_cell_two = data_from_cell_two.lower()
            result_search = re.findall(search_text, data_from_cell_two)
            if len(result_search) > 0:
                result_row = i + 1
                break
            elif data_from_cell is None:
                result_row = i + 1
                break

        filling_restr_two(self, result_row, full_name)

    except (Exception,):
        mistake = 'Неправельный путь к Реестру ЭП'
        exception(mistake)


def filling_restr_two(self, result_row, full_name):
    try:
        reestr = 'Реестр ЭП'
        config = configparser.ConfigParser()
        config.read("doc\\settings.ini", encoding="utf-8")
        date_of_completion = datetime.datetime.today().strftime('%d.%m.%Y')
        date_expiration = datetime.datetime.today().strftime('%d.%m.2023')
        ser_num = self.ui.ser_num.text()
        wb = openpyxl.load_workbook(config["Path"]["path_reestr_es_full"], read_only=False)
        sheet = wb.get_sheet_by_name('Реестр выпущенных ЭП')

        sheet.cell(row=result_row, column=1).value = full_name
        sheet.cell(row=result_row, column=4).value = ser_num
        sheet.cell(row=result_row, column=5).value = date_of_completion
        sheet.cell(row=result_row, column=6).value = date_expiration
        sheet.cell(row=result_row, column=7).value = 'ООО ИК "СИБИНТЕК"'
        sheet.cell(row=result_row, column=8).value = 'Обычная'
        sheet.cell(row=result_row, column=9).value = 'План выпуска ЭП (1200)'
        sheet.cell(row=result_row, column=10).value = 'Действует'
        sheet.cell(row=result_row, column=11).value = 'В целях автоматизации ЭД'
        sheet.cell(row=result_row, column=13).value = date_of_completion
        sheet.cell(row=result_row, column=14).value = date_of_completion
        sheet.cell(row=result_row, column=15).value = date_of_completion

        wb.save(config["Path"]["path_reestr_es_full"])
        successfully(result_row, full_name, reestr)
    except (Exception,):
        mistake = 'Не удалось отметить в реестре ЭЦП\nПроверьте, что файл не кто не использует!'
        exception(mistake)


def successfully(result_row, full_name, reestr):
    date_of_completion = datetime.datetime.today().strftime('%d.%m.%Y')
    suc = QMessageBox()
    suc.setWindowIcon(QIcon("image\\186100_900.jpg"))
    suc.setWindowTitle("Успех")
    suc.setText(f"{reestr}\nПользователь: {full_name}\nСтрока: {result_row}\nДата: {date_of_completion}")
    suc.setIcon(QMessageBox.Information)
    suc.setStandardButtons(QMessageBox.Ok)
    suc.exec_()


def exception(mistake):
    error = QMessageBox()
    error.setWindowIcon(QIcon("image\\186100_900.jpg"))
    error.setWindowTitle("Ошибка")
    error.setText('В работе возникла ошибка')
    error.setIcon(QMessageBox.Warning)
    error.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
    error.setInformativeText(mistake)
    error.exec_()


# функции второго окна

def start_config(self):
    config = configparser.RawConfigParser()
    config.read("doc\\settings.ini", encoding="utf-8")
    self.ui.line_page_page.insert(config["Page"]["page"])
    self.ui.line_page_row.insert(config["Page"]["row"])

    self.ui.line_path_path_db_all.insert(config["Path"]["path_db_all"])
    self.ui.line_path_path_reestr_es.insert(config["Path"]["path_reestr_es"])
    self.ui.line_path_path_reestr_es_full.insert(config["Path"]["path_reestr_es_full"])
    self.ui.line_path_path_pkzi.insert(config["Path"]["path_pkzi"])
    self.ui.line_path_path_for_user.insert(config["Path"]["path_for_user"])
    self.ui.line_path_path_for_archives.insert(config["Path"]["path_for_archives"])
    self.ui.line_path_path_downloads.insert(config["Path"]["path_downloads"])
    self.ui.line_path_path_for_receipt.insert(config["Path"]["path_for_receipt"])

    self.ui.line_url_url.insert(config["URL"]["url"])
    self.ui.line_url_login.insert(config["URL"]["login"])
    self.ui.line_url_password.insert(config["URL"]["pass"])


def save_config(self):
    config = configparser.RawConfigParser()
    config.add_section('Page')
    config.set('Page', 'page', self.ui.line_page_page.text())
    config.set('Page', 'row', self.ui.line_page_row.text())

    config.add_section('Path')
    config.set('Path', 'path_db_all', self.ui.line_path_path_db_all.text())
    config.set('Path', 'path_reestr_es', self.ui.line_path_path_reestr_es.text())
    config.set('Path', 'path_reestr_es_full', self.ui.line_path_path_reestr_es_full.text())
    config.set('Path', 'path_pkzi', self.ui.line_path_path_pkzi.text())
    config.set('Path', 'path_for_user', self.ui.line_path_path_for_user.text())
    config.set('Path', 'path_for_archives', self.ui.line_path_path_for_archives.text())
    config.set('Path', 'path_downloads', self.ui.line_path_path_downloads.text())
    config.set('Path', 'path_for_receipt', self.ui.line_path_path_for_receipt.text())

    config.add_section('URL')
    config.set('URL', 'url', self.ui.line_url_url.text())
    config.set('URL', 'login', self.ui.line_url_login.text())
    config.set('URL', 'pass', self.ui.line_url_password.text())

    with open("doc\\settings.ini", 'w', encoding="utf-8") as f:
        config.write(f)

    successfully_config()


def successfully_config():
    suc = QMessageBox()
    suc.setWindowIcon(QIcon("image\\186100_900.jpg"))
    suc.setWindowTitle("Успех")
    suc.setText(f"Файл settings.ini успешно изменен")
    suc.setIcon(QMessageBox.Information)
    suc.setStandardButtons(QMessageBox.Ok)
    suc.exec_()
